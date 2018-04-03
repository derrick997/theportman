from termcolor import colored
import csv
from Stock import Stock
import FinancialsCatcher
import StockCatcher
from openpyxl import Workbook
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
import datetime


def cumulativeHoldingAnalysisToString(stock_list):
    
    print("-----Start of summary-----")
    
    total_paid = 0.00
    total_current = 0.00
    total_gain = 0.00

    for stock in stock_list:
        print(stock.quoteToString())

        total_paid += (stock.numberbought * stock.pricebought)
        total_current += (stock.numberbought * stock.price)

        gainloss = stock.price - stock.pricebought
        total_gain += gainloss

        if gainloss < 0:
            gainloss *= -1
            print(colored("Loss per share so far: $(" + str(gainloss) + ")", "red"))
            print(colored("Total loss so far: $(" + str(gainloss * stock.numberbought) + ")", "red") + "\n")
        else:
            print(colored("Gain per share so far: $" + str(gainloss), "green"))
            print(colored("Total gain so far: $" + str(gainloss * stock.numberbought), "green") + "\n")

    print("Total portfolio cost: $" + str(total_paid))
    print("Total portfolio value: $" + str(total_current))
    if total_gain < 0:
        print(colored("Total current loss: $(" + str(total_gain * -1) + ")", "red") + "\n")
    else:
        print(colored("Total current gain: $" + str(total_gain), "green") + "\n")

    print("-----End of summary-----")

def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        return False

def listOfListsToXlsx(worksheet, datalist):
    index = 1
    for sublist in datalist:
        subIndex = 1
        for element in sublist:
            address = chr(subIndex + 64) + str(index)
            # print(address)
            worksheet[address] = element
            #print(type(element))
            '''if (type(element) == float) or (type(element) == int):
                worksheet[address].number_format'''
            subIndex += 1
        index += 1
    return worksheet

def savePortfolioAsXlsx(stock_list, bond_list, wacc_list):

    save = None
    while (save != "Y" and save != "y" and save != "N" and save != "n"):
        save = input("Save this portfolio as xlsx? (Y/N): ")
        save = save.strip()
    if (save == "N" or save == "n"):
        print("This portfolio will not be saved.")
    else:
        for stock in stock_list:
            filename = str(stock.getName()) + ".xlsx"
            wb = Workbook()
            wb.guess_types = True

            worksheet_list = []

            ws0 = wb.active
            ws0.title="Quote Analysis Summary"
            ws0 = listOfListsToXlsx(ws0, stock.quoteToListOfLists())
            #print(stock.quoteToListOfLists)
            worksheet_list.append(ws0)

            if (stock.getHoldingAnalysisList() != []):
                ws1 = wb.create_sheet(title="Stock Holding Analysis")
                ws1 = listOfListsToXlsx(ws1, stock.getHoldingAnalysisList())
                worksheet_list.append(ws1)

            if (stock.getIS_numYearsAvailable() > 0):
                ws2 = wb.create_sheet(title="Historical Income Statement")
                ISList = stock.incomeStatementToListOfLists()
                ws2 = listOfListsToXlsx(ws2, ISList)
                worksheet_list.append(ws2)

            if (stock.getBS_numYearsAvailable() > 0):

                ws3 = wb.create_sheet(title="Historical Balance Sheet")
                BSList = stock.balanceSheetToListOfLists()
                ws3 = listOfListsToXlsx(ws3, BSList)
                worksheet_list.append(ws3)

            if (stock.getCF_numYearsAvailable() > 0):

                CFList = stock.cashFlowStatementToListOfLists()
                ws4 = wb.create_sheet(title="Historical Cash Flow Statement")
                ws4 = listOfListsToXlsx(ws4, CFList)
                worksheet_list.append(ws4)

            if (bond_list != None):

                ws5 = wb.create_sheet(title="Treasury Bonds")
                ws5 = listOfListsToXlsx(ws5, bond_list)
                worksheet_list.append(ws5)

            if (wacc_list != None):

                ws6 = wb.create_sheet(title="WACC")
                ws6 = listOfListsToXlsx(ws6, wacc_list)
                worksheet_list.append(ws6)

            # Auto adjust column widths
            for worksheet in worksheet_list:
                for column_cells in worksheet.columns:
                    length = max(len(str(cell.value)) for cell in column_cells)
                    worksheet.column_dimensions[column_cells[0].column].width = length

            # Bold titles
            '''boldList = ['A1', 'B3', 'C3', 'D3', 'A5', 'A11', 'A20', 'A30', 'A39']
            for cellAddr in boldList:
                _cell = ws2`.cell(cellAddr)
                _cell.style.font.bold = True'''

            wb.save(filename=filename)

            print("Your portfolio has been saved as " + filename)

#Test for save xlsx
'''stockList = []
dominos = StockCatcher.getStockFromYahooFinance("MyDominosStocks", "DPZ", Stock())

dominos = FinancialsCatcher.catchIncomeStatementFromYahooFinance(dominos)
dominos = FinancialsCatcher.catchBalanceSheetFromYahooFinance(dominos)

stockList.append(dominos)
savePortfolioAsXlsx(stockList)'''


def savePortfolioAsCsv(stock_list):

    save = None
    while (save != "Y" and save != "y" and save != "N" and save != "n"):
        save = input("Save this portfolio as csv? (Y/N): ")
        save = save.strip()
    if (save == "N" or save == "n"):
        print("This portfolio will not be saved.")
    else:
        for stock in stock_list:
            #print(stock.incomeStatementToListOfLists())
            filename = str(stock.getName()) + ".csv"
            with open(filename, 'w') as portFile:
                writer = csv.writer(portFile)
                writer.writerows(stock.incomeStatementToListOfLists())
            portFile.close()
            print("Your portfolio has been saved as " + colored(filename, "yellow"))

        # Add file name
        ''' name_confirmation = None
            first_try = True
            while (name_confirmation != "Y" and name_confirmation != "y"):

            if first_try:
                name = input("Enter portfolio name: ")
                first_try = False
            else:
                name = input("Please re-enter portfolio name: ")
            name = name.strip()
            print(name)

            second_confirmation = None
            while (second_confirmation != "Y" and second_confirmation != "y" and second_confirmation != "N" and second_confirmation != "n"):
                second_confirmation = input('Is "' + name + '" correct? (Y/N): ')
                second_confirmation = second_confirmation.strip()
            if (second_confirmation == "N" or second_confirmation == "n"):
                print("This portfolio will not be saved.")
            else:
                with open(name + '.csv', 'wb') as portFile:
                    writer = csv.writer(portFile)
                    
                    for stock in stock_list:
                        writer.writerows(stock.incomeStatementToListOfLists())

                portFile.close()'''

#Test for save csv
'''stockList = []
dominos = StockCatcher.getStockFromYahooFinance("MyDominosStocks", "DPZ", Stock())
#print(dominos.quoteToString())
dominos = FinancialsCatcher.getFinancialsFromYahooFinance(dominos)
#print(dominos.financialToString())
stockList.append(dominos)
savePortfolio(stockList)'''